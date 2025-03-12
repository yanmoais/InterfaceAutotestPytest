"""
    读取Excel文件的方法
"""
from openpyxl import load_workbook
from config.Base_Env import *
from util_tools.logger import Logger


class read_excel():
    def __init__(self, excle_file=CASE_FILE_PATH, sheet_name=SHEET_NAME_PATH):
        self.excel_obj = load_workbook(excle_file)
        self.sheet_obj = self.excel_obj[sheet_name]
        self.logging = Logger().init_logger()

    def _get_title_and_data(self):
        case_datas = list(self.sheet_obj.iter_rows(values_only=True))
        case_title, case_data = case_datas[0], case_datas[1:]
        return case_title, case_data

    def _close_excel(self):
        self.excel_obj.close()

    def write_save_excel(self, case_name, status):
        # 案例名称在G列，所以得从案例名称开始找
        for row in self.sheet_obj.iter_rows(min_row=2):  # 从第二行开始找
            if row[6].value == case_name:
                row[7].value = status
                break
        try:
            self.excel_obj.save(CASE_FILE_PATH)
        except FileNotFoundError:
            self.logging.error(f"保存出错，未找到该文件 {CASE_FILE_PATH}！请检查是否误删或者改名！")
        finally:
            self.excel_obj.close()

    def get_case_data(self):
        case_data_list = []
        case_title, case_data = self._get_title_and_data()
        for result in case_data:
            json_data = dict(zip(case_title, result))
            case_data_list.append(json_data)
        return case_data_list

    def get_case_private_params(self, case_data, keys):
        case_param_list = []
        for results in case_data:
            case_param_list.append(results[keys])
        return case_param_list


if __name__ == '__main__':
    term = '9'
    if term == '9':
        sheet_obj = read_excel(sheet_name='Sheet_9')
        print(sheet_obj.get_case_data())
